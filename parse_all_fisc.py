import re,os,unicodedata,numpy as np,pandas as pd
from openpyxl import load_workbook
def norm(s):
    s=unicodedata.normalize('NFKD',str(s)).encode('ascii','ignore').decode().lower()
    return re.sub(r'[^a-z0-9 ]',' ',re.sub(r'\s+',' ',s)).strip()
M3={'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}
MFULL={'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,'julio':7,'agosto':8,
       'septiembre':9,'setiembre':9,'octubre':10,'noviembre':11,'diciembre':12}
def periodo(fn):
    f=norm(fn.replace('.xlsx',''))
    m=re.search(r'([a-z]+)[ _]*(\d{2,4})',f)
    if not m: return None
    nm,yr=m.group(1),m.group(2)
    mo=M3.get(nm[:3]) if nm[:3] in M3 else MFULL.get(nm)
    if nm in MFULL: mo=MFULL[nm]
    if not mo: return None
    y=int(yr); y=2000+y if y<100 else y
    if not (2004<=y<=2026): return None
    return pd.Period(f"{y}-{mo:02d}",freq='M')

PAT={'gasto_primario':lambda k:'gastos primarios despues' in k or k=='gastos primarios',
     'resultado_primario':lambda k:('superavit primario' in k or k=='resultado primario') and 'sin privatizac' not in k and 'facili' not in k,
     'resultado_financiero':lambda k:'resultado financiero' in k and 'sin privatizac' not in k and 'facili' not in k,
     'ingresos_totales':lambda k:k=='ingresos totales'}

def leer(path):
    wb=load_workbook(path,read_only=True,data_only=True)
    hojas=[h for h in wb.sheetnames if h.upper() not in ('VARMENSUAL','MENSUALIZACION','MENSUALIZACIÓN') and 'PRENSA' not in h.upper()]
    hojas=hojas or [wb.sheetnames[0]]
    for h in hojas:
        ws=wb[h]; rec={}
        for r in ws.iter_rows(max_col=14,values_only=True):
            strs=[x for x in r if isinstance(x,str) and len(x.strip())>5]
            nums=[x for x in r if isinstance(x,(int,float)) and not isinstance(x,bool)]
            if not strs or not nums: continue
            k=norm(strs[0])
            for name,test in PAT.items():
                if test(k) and name not in rec:
                    # esquema AIF: columnas por jurisdiccion, el TOTAL es la ultima
                    # esquema IMIG: una sola columna de dato mensual, es la primera
                    aif_style=('xi' in k.split()) or k.startswith('superavit primario') or 'despues de figurat' in k or 'despues figurat' in k
                    rec[name]=float(nums[-1] if aif_style else nums[0])
        if 'gasto_primario' in rec and 'resultado_primario' in rec: return rec
    return rec

out={}
for src in ['/tmp/f2','/tmp/f3']:
    for f in sorted(os.listdir(src)):
        if not f.endswith('.xlsx'): continue
        p=periodo(f)
        if p is None: print('sin periodo:',f); continue
        try: rec=leer(os.path.join(src,f))
        except Exception as e: print('error',f,e); continue
        if rec: out.setdefault(p,rec)
per=sorted(out)
df=pd.DataFrame([out[p] for p in per],index=pd.PeriodIndex([str(p) for p in per],freq='M')).sort_index()
print('meses:',len(df),df.index.min(),'->',df.index.max())
falt=pd.period_range('2004-01','2026-07',freq='M').difference(df.index)
print('faltantes:',len(falt),list(falt)[:20])
print('NaN:',df.isna().sum().to_dict())
df.to_csv('fiscal_mensual_nominal.csv')
