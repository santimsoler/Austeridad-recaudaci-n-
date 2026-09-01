import numpy as np,pandas as pd,re
from openpyxl import load_workbook
E='est/'
def rows(f,sh=None,c=8):
    wb=load_workbook(E+f,read_only=True,data_only=True)
    ws=wb[sh] if sh else wb[wb.sheetnames[0]]
    return [r for r in ws.iter_rows(max_col=c,values_only=True) if any(x is not None for x in r)]

# --- IPC (variacion mensual) ---
ipc={}
for r in rows('ipc mensual (1).xlsx',c=2):
    if hasattr(r[0],'year'): ipc[(r[0].year,r[0].month)]=float(r[1])
# meses faltantes 2026 (INDEC, informes de prensa)
for m,v in {4:2.6,5:2.1,6:1.9,7:2.1}.items(): ipc.setdefault((2026,m),v)
ipc=pd.Series(ipc); ipc.index=pd.PeriodIndex([f"{a}-{b:02d}" for a,b in ipc.index],freq='M')
ipc=ipc.sort_index()
nivel=(1+ipc/100).cumprod(); nivel/=nivel.loc['2026-07']   # base jul-26 = 1

# --- EMAE nivel general ---
meses=['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
emae={};emae_sa={};y=None
for r in rows('sh_emae_mensual_base2004__1_.xlsx',c=7):
    if isinstance(r[0],(int,float)) and r[0] and 1990<r[0]<2100: y=int(r[0])
    elif isinstance(r[0],str) and r[0].strip().isdigit(): y=int(r[0])
    if r[1] in meses and y:
        p=pd.Period(f"{y}-{meses.index(r[1])+1:02d}",freq='M')
        emae[p]=float(r[2]); emae_sa[p]=float(r[4])
emae=pd.Series(emae).sort_index(); emae_sa=pd.Series(emae_sa).sort_index()

# --- Supermercados precios constantes ---
sup={}
for r in rows('serie_supermercados.xlsx','Cuadro 3.',c=3):
    if hasattr(r[0],'year'): sup[pd.Period(f"{r[0].year}-{r[0].month:02d}",freq='M')]=float(r[2])
sup=pd.Series(sup).sort_index()

# --- ITCRM ---
itc={}
for r in rows('ITCRMSerie (1).xlsx','ITCRM y bilaterales prom. mens.',c=2):
    if hasattr(r[0],'year'):
        try: itc[pd.Period(f"{r[0].year}-{r[0].month:02d}",freq='M')]=float(r[1])
        except: pass
itc=pd.Series(itc).sort_index()

# --- Empleo registrado (SIPA) ---
emp={}
MM={'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}
for r in rows('trabajoregistrado_2605_estadisticas.xlsx','T.1',c=4):
    p=None
    if hasattr(r[0],'year'): p=pd.Period(f"{r[0].year}-{r[0].month:02d}",freq='M')
    elif isinstance(r[0],str):
        m=re.match(r'([a-z]{3})-(\d{2})',r[0].strip().lower())
        if m and m.group(1) in MM: p=pd.Period(f"20{m.group(2)}-{MM[m.group(1)]:02d}",freq='M')
    if p is not None:
        try: emp[p]=float(r[2])
        except: pass
emp=pd.Series(emp).sort_index()

# --- RIPTE: variaciones -> nivel encadenado ---
ws=load_workbook(E+'ripte.xlsx',read_only=True,data_only=True)['Hoja1']
rr=[r for r in ws.iter_rows(max_col=14,values_only=True)]
hdr=[i for i,r in enumerate(rr) if r[1]=='ENE'][0]
rip={}
for r in rr[hdr+1:]:
    try: y=int(r[0])
    except: continue
    for m in range(1,13):
        v=r[m+1]
        if v is not None:
            try: rip[pd.Period(f"{y}-{m:02d}",freq='M')]=float(v)
            except: pass
rip=pd.Series(rip).sort_index()
ripte_nivel=(1+rip).cumprod()

# --- Consumo privado trimestral (desest) ---
cons={}
y=None
for r in rows('sh_oferta_demanda_desest_06_26.xlsx','desestacionalizado n',c=7):
    if isinstance(r[0],(int,float)) and r[0] and 1990<r[0]<2100: y=int(r[0])
    if r[1] in ['I','II','III','IV'] and y:
        cons[pd.Period(f"{y}Q{['I','II','III','IV'].index(r[1])+1}",freq='Q')]=float(r[4])
cons=pd.Series(cons).sort_index()

def cob(s,n):
    print(f"{n:26} {str(s.index.min()):>9} -> {str(s.index.max()):>9}  n={len(s)}")
for s,n in [(ipc,'IPC var% mensual'),(emae,'EMAE nivel general'),(emae_sa,'EMAE desest.'),
            (sup,'Supermercados const.'),(itc,'ITCRM'),(emp,'Empleo registrado'),
            (ripte_nivel,'RIPTE nivel'),(cons,'Consumo privado (trim)')]:
    cob(s,n)

df=pd.DataFrame({'ipc_var':ipc,'defl':nivel,'emae':emae,'emae_sa':emae_sa,
                 'supermercados':sup,'itcrm':itc,'empleo':emp,'ripte':ripte_nivel})
df=df[df.index>=pd.Period('2010-01',freq='M')]
df.to_csv('panel_mensual.csv'); cons.to_csv('consumo_trim.csv')
print("\nsolape completo mensual:", df.dropna().index.min(),'->',df.dropna().index.max(), len(df.dropna()))
