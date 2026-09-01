import re,os,unicodedata,numpy as np,pandas as pd
from openpyxl import load_workbook
def norm(s):
    s=unicodedata.normalize('NFKD',str(s)).encode('ascii','ignore').decode().lower()
    return re.sub(r'[^a-z0-9 ]',' ',re.sub(r'\s+',' ',s)).strip()
M3={'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}
MF={'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,'julio':7,'agosto':8,
    'septiembre':9,'setiembre':9,'octubre':10,'noviembre':11,'diciembre':12}
def periodo(fn):
    f=norm(fn.replace('.xlsx',''))
    m=re.search(r'([a-z]+)[ _]*(\d{2,4})',f)
    if not m: return None
    nm,yr=m.group(1),m.group(2)
    mo=MF.get(nm) or M3.get(nm[:3])
    if not mo: return None
    y=int(yr); y=2000+y if y<100 else y
    return pd.Period(f"{y}-{mo:02d}",freq='M') if 2004<=y<=2026 else None

def aif_rows(ws):
    r_={}
    for r in ws.iter_rows(max_col=14,values_only=True):
        strs=[x for x in r if isinstance(x,str) and len(x.strip())>4]
        nums=[x for x in r if isinstance(x,(int,float)) and not isinstance(x,bool)]
        if not strs or not nums: continue
        k=norm(strs[0])
        if ('gastos antes' in k and 'figurat' in k) and 'gasto_antes' not in r_:
            r_['gasto_antes']=float(nums[-1])
        if k.startswith('ii ') or k=='gastos corrientes':
            r_.setdefault('gastos_corrientes',float(nums[-1]))
        if 'gastos de capital' in k: r_.setdefault('gastos_capital',float(nums[-1]))
        if ('interes' in k) and ('intra' not in k) and ('pagados' not in k) and ('excluye' not in k) and ('otras rentas' not in k):
            r_.setdefault('intereses',float(nums[-1]))
    return r_
out={}
for src in ['/tmp/f2','/tmp/f3']:
    for f in sorted(os.listdir(src)):
        if not f.endswith('.xlsx'): continue
        p=periodo(f)
        if p is None or p in out: continue
        try: wb=load_workbook(os.path.join(src,f),read_only=True,data_only=True)
        except: continue
        best={}
        for h in wb.sheetnames:
            if h.upper() in ('VARMENSUAL','MENSUALIZACION','MENSUALIZACIÓN') or 'PRENSA' in h.upper(): continue
            r=aif_rows(wb[h])
            if 'gasto_antes' in r: best=r; break
            if r and not best: best=r
        if best: out[p]=best
df=pd.DataFrame([out[p] for p in sorted(out)],index=pd.PeriodIndex([str(p) for p in sorted(out)],freq='M'))
print('meses:',len(df),'| sin gasto_antes:',int(df['gasto_antes'].isna().sum()) if 'gasto_antes' in df else 'n/a')
print('faltan:',[str(x) for x in df[df['gasto_antes'].isna()].index][:20])
df.to_csv('gasto_antes.csv')
