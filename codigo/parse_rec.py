import re,os,unicodedata
import pandas as pd,numpy as np
from openpyxl import load_workbook
MES={'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,'julio':7,
     'agosto':8,'septiembre':9,'setiembre':10 if False else 9,'octubre':10,'noviembre':11,'diciembre':12}
def norm(s):
    s=unicodedata.normalize('NFKD',str(s)).encode('ascii','ignore').decode().lower()
    return re.sub(r'\s+',' ',s).strip(' .:-')
rows={}
bad=[]
for f in sorted(os.listdir('/tmp/conv')):
    if not f.endswith('.xlsx'): continue
    m=re.search(r'(20\d{2})[-_]+([a-zA-Zé]+)',f)
    if not m: bad.append(f); continue
    y=int(m.group(1)); mn=MES.get(norm(m.group(2)))
    if not mn: bad.append(f); continue
    try:
        ws=load_workbook('/tmp/conv/'+f,read_only=True,data_only=True).active
        data=[r for r in ws.iter_rows(max_col=4,values_only=True)]
    except Exception as e:
        bad.append(f); continue
    d={}
    # detectar la columna donde esta 'Concepto'
    ci=0
    for r in data:
        for j in range(3):
            if r[j] and str(r[j]).strip()=='Concepto': ci=j
    for r in data:
        if r[ci] and isinstance(r[ci+1],(int,float)):
            d[norm(r[ci])]=float(r[ci+1])
    if d: rows[(y,mn)]=d
    else: bad.append(f)
print('parseados:',len(rows),'| fallidos:',len(bad))
if bad: print('fallidos:',bad[:10])
per=sorted(rows)
print('rango:',per[0],'->',per[-1])
# conceptos presentes en >90% de los meses
from collections import Counter
c=Counter(k for d in rows.values() for k in d)
keep=[k for k,v in c.items() if v>len(rows)*0.85]
print('conceptos estables:',len(keep))
df=pd.DataFrame([{**{'periodo':pd.Period(f"{y}-{m:02d}",freq='M')},
                  **{k:rows[(y,m)].get(k,np.nan) for k in keep}} for y,m in per]).set_index('periodo')
df.to_csv('rec_larga_nominal.csv')
for k in sorted(keep)[:60]: print(' -',k)

# ---- backfill de meses faltantes usando la columna "mes anterior" / "mismo mes año previo" ----
