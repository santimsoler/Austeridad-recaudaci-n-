import matplotlib; matplotlib.use('Agg')
import matplotlib.pyplot as plt, pandas as pd, numpy as np
from openpyxl import load_workbook
plt.rcParams.update({'font.size':9,'axes.spines.top':False,'axes.spines.right':False,'figure.dpi':200})
AZ='#1b4f72'; NA='#e67e22'; VD='#1e8449'
ws=load_workbook('est/sh_emae_actividad_base2004.xlsx',read_only=True,data_only=True)['Tabla Letras']
rows=[r for r in ws.iter_rows(max_col=20,values_only=True)]; hdr=rows[2]
meses=['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
data={};y=None
for r in rows[5:]:
    if r[0] not in (None,''):
        try: y=int(float(r[0]))
        except: pass
    if r[1] in meses and y:
        data[pd.Period(f"{y}-{meses.index(r[1])+1:02d}",freq='M')]={hdr[j]:r[j] for j in range(2,len(hdr)) if hdr[j] and isinstance(r[j],(int,float))}
df=pd.DataFrame(data).T.sort_index()
sel={c:('Industria' if 'Industria manufacturera' in c else 'Construcción' if 'Construcción' in c else 'Comercio')
     for c in df.columns if any(k in c for k in ['Industria manufacturera','Construcción','Comercio mayorista'])}
d=df[list(sel)].rename(columns=sel)
ia={'Industria':1.019,'Construcción':1.028,'Comercio':1.008}
jun=pd.Series({k:d.loc['2025-06',k]*v for k,v in ia.items()},name=pd.Period('2026-06',freq='M'))
d=pd.concat([d,jun.to_frame().T])
mm=d.rolling(12).mean().loc['2024-12':]
mm=mm/mm.iloc[0]*100
fig,ax=plt.subplots(figsize=(7.2,3.4))
for c,col in [('Industria',AZ),('Construcción',NA),('Comercio',VD)]:
    ax.plot(range(len(mm)),mm[c],color=col,lw=2.2,label=c)
j=list(mm.index).index(pd.Period('2025-07',freq='M'))
ax.axvline(j,color='#999',ls='--',lw=1)
ax.annotate('julio 2025:\nse corta\nla recuperación',xy=(j,mm['Comercio'].iloc[j]),xytext=(j-3.6,105.6),
            fontsize=8.2,color='#555',ha='center',arrowprops=dict(arrowstyle='->',color='#999',lw=.9))
ax.annotate('junio 2026:\nlos tres vuelven\na crecer',xy=(len(mm)-1,mm['Construcción'].iloc[-1]),xytext=(len(mm)-6.2,100.6),
            fontsize=8.2,color='#333',ha='center',arrowprops=dict(arrowstyle='->',color='#666',lw=.9))
tk=[i for i,p in enumerate(mm.index) if p.month in (1,7)]
ax.set_xticks(tk); ax.set_xticklabels([f"{'ene' if mm.index[i].month==1 else 'jul'}-{str(mm.index[i].year)[2:]}" for i in tk],fontsize=8)
ax.set_ylabel('Tendencia (dic-24 = 100)')
ax.set_ylim(99,107)
ax.set_title('Se recuperaban, se frenaron, y vuelven a moverse',fontweight='bold',loc='left')
ax.legend(frameon=False,ncol=3,loc='lower center',bbox_to_anchor=(.5,-.26),fontsize=8.5)
ax.text(.99,-.32,'Media móvil de 12 meses. Junio 2026 estimado con la variación interanual publicada.',
        transform=ax.transAxes,ha='right',fontsize=7,color='#888')
fig.tight_layout(); fig.savefig('fig/g7.png',bbox_inches='tight'); plt.close()
print('ok')
