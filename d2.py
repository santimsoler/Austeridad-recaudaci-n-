from openpyxl import load_workbook
meses=['ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']
rec={}
for y in range(2019,2027):
    wb=load_workbook(f'conv/serie{y}.xlsx',read_only=True,data_only=True)
    ws=wb[wb.sheetnames[0]]
    rows=list(ws.iter_rows(max_row=12,max_col=18,values_only=True))
    hdr=tot=None
    for r in rows:
        if r[1]=='CONCEPTO': hdr=r
        if r[1] and 'TOTAL GENERAL' in str(r[1]): tot=r
    for i,h in enumerate(hdr):
        if h in meses and tot[i] is not None:
            rec[(y,meses.index(h)+1)]=tot[i]

ipc_y={
2019:[2.9,3.8,4.7,3.4,3.1,2.7,2.2,4.0,5.9,3.3,4.3,3.7],
2020:[2.3,2.0,3.3,1.5,1.5,2.2,1.9,2.7,2.8,3.8,3.2,4.0],
2021:[4.0,3.6,4.8,4.1,3.3,3.2,3.0,2.5,3.5,3.5,2.5,3.8],
2022:[3.9,4.7,6.7,6.0,5.1,5.3,7.4,7.0,6.2,6.3,4.9,5.1],
2023:[6.0,6.6,7.7,8.4,7.8,6.0,6.3,12.4,12.7,8.3,12.8,25.5],
2024:[20.6,13.2,11.0,8.8,4.2,4.6,4.0,4.2,3.5,2.7,2.4,2.7],
2025:[2.2,2.4,3.7,2.8,1.5,1.6,1.9,1.9,2.1,2.3,2.5,2.8],
2026:[2.9,2.9,3.4,2.6,2.1,1.9,2.1],
}
ipc={(y,m+1):v for y,l in ipc_y.items() for m,v in enumerate(l)}
per=sorted(rec)
idx={};lvl=100.0
# arrancar antes de ene2019
for p in per:
    lvl*=(1+ipc[p]/100); idx[p]=lvl
for y in range(2019,2026):
    print(y,'acum %',round((idx[(y,12)]/idx[(y-1,12)]-1)*100,1) if (y-1,12) in idx else '-')
base=idx[(2026,7)]
d={}
print(f"\n{'Mes':9}{'Nominal':>10}{'Real':>10}{'i.a.':>9}")
for p in per:
    f=1e9 if p[0]<=2020 else 1e6
    n=rec[p]/f; r=n*base/idx[p]; d[p]=r
    prev=(p[0]-1,p[1]); ia=f"{(r/d[prev]-1)*100:+.1f}%" if prev in d else ""
    print(f"{p[1]:02d}/{p[0]} {n:9.2f} {r:9.2f} {ia:>9}")
print()
for y in range(2019,2027):
    tot=sum(d[p] for p in per if p[0]==y)
    n=len([p for p in per if p[0]==y])
    print(y,'total real',round(tot,1),'prom mensual',round(tot/n,2),f'({n} meses)')
